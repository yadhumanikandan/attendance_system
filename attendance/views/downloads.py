"""
Download views for generating XLSX reports.
Handles monthly report downloads for both in-house and remote employees.
"""

import datetime
import calendar
from datetime import time
from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..models import (
    Employee, AttendanceRecord, MonthlySummary, ShiftHistory, Holiday,
    RemoteEmployee, RemoteCallRecord, RemoteMonthlySummary, LeaveRequest
)


@login_required
def download_report(request):
    """Generate and download XLSX report for the selected month."""
    now = datetime.datetime.now()
    try:
        selected_month = int(request.GET.get('month', now.month))
        selected_year = int(request.GET.get('year', now.year))
    except ValueError:
        selected_month = now.month
        selected_year = now.year
    
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[selected_month]
    
    _, days_in_month = calendar.monthrange(selected_year, selected_month)
    sundays_count = 0
    for day in range(1, days_in_month + 1):
        date = datetime.date(selected_year, selected_month, day)
        if date.weekday() == 6:
            sundays_count += 1
    
    show_inactive = request.GET.get('show_inactive', '') == '1'
    
    summaries = MonthlySummary.objects.filter(
        year=selected_year,
        month=selected_month
    ).select_related('employee')
    
    if not show_inactive:
        summaries = summaries.filter(employee__is_active=True)
    
    summaries = summaries.order_by('employee__name')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {selected_year}"
    
    # Styles
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:G1')
    ws['A1'] = f"Attendance Report - {month_name} {selected_year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.append([])
    
    holidays_in_month = Holiday.objects.filter(
        date__year=selected_year,
        date__month=selected_month
    ).count()
    total_holidays = sundays_count + holidays_in_month
    
    headers = ['Employee Name', 'Full Days', 'Half Days', 'Paid Leave', 'Leave Days', 'Late Arrivals', 'Holidays', 'Total Working Days']
    ws.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for summary in summaries:
        # Calculate paid leave days
        employee = summary.employee
        month_start = datetime.date(selected_year, selected_month, 1)
        month_end = datetime.date(selected_year, selected_month, days_in_month)
        
        approved_leaves = LeaveRequest.objects.filter(
            employee=employee,
            status='approved',
            start_date__lte=month_end,
            end_date__gte=month_start
        )
        approved_leave_days_count = 0
        approved_leave_dates = set()
        
        for leave in approved_leaves:
            start = max(leave.start_date, month_start)
            end = min(leave.end_date, month_end)
            curr = start
            while curr <= end:
                if curr.weekday() != 6: # Exclude Sundays from paid leave count if that's the rule, but usually paid leave counts as day.
                    # Actually, if it's a holiday, it's a holiday. Paid leave usually overlaps working days.
                    # Let's count all days in range that are not Sundays (assuming Sundays are weekly off).
                    # But need to check if holiday.
                     approved_leave_dates.add(curr)
                curr += datetime.timedelta(days=1)
                
        # Filter out Sundays and Holidays from paid leave count
        holidays_qs = Holiday.objects.filter(date__range=(month_start, month_end))
        holiday_date_set = set(h.date for h in holidays_qs)
        
        real_paid_leave_count = 0
        for d in approved_leave_dates:
            if d.weekday() != 6 and d not in holiday_date_set:
                real_paid_leave_count += 1
        
        half_days = getattr(summary, 'half_days', 0) or 0
        full_days = max(0, summary.working_days - half_days)
        # leave_days in MonthlySummary might include paid leaves depending on how it was updated.
        # In reports.py we did: leave_days = expected - actual.
        # Then we subtracted paid_leave_count. 
        # But MonthlySummary stored `leave_days` (after subtraction). 
        # So summary.leave_days should be correct (unpaid).
        
        leave_days = max(0, summary.leave_days - real_paid_leave_count)
        working_days_total = full_days + (half_days * 0.5) + total_holidays + real_paid_leave_count
        
        row = [
            summary.employee.name,
            full_days,
            half_days,
            real_paid_leave_count,
            leave_days,
            summary.late_days,
            total_holidays,
            working_days_total
        ]
        ws.append(row)
        
        current_row = ws.max_row
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            if col > 1:
                cell.alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Attendance_Report_{selected_year}_{selected_month:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Content-Length'] = len(buffer.getvalue())
    
    return response


@login_required
def download_employee_report(request, employee_id):
    """Generate and download XLSX report for a single employee for the selected month."""
    employee = get_object_or_404(Employee, id=employee_id)
    
    now = datetime.datetime.now()
    try:
        selected_month = int(request.GET.get('month', now.month))
        selected_year = int(request.GET.get('year', now.year))
    except ValueError:
        selected_month = now.month
        selected_year = now.year
    
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[selected_month]
    
    holiday_dates = set(Holiday.objects.filter(
        date__year=selected_year,
        date__month=selected_month
    ).values_list('date', flat=True))
    
    records = AttendanceRecord.objects.filter(
        employee=employee,
        date__year=selected_year,
        date__month=selected_month
    ).order_by('date')
    
    records_dict = {r.date: r for r in records}
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{employee.name[:20]}"
    
    # Styles
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    holiday_fill = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")
    sunday_fill = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    red_fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Attendance Report - {month_name} {selected_year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Employee: {employee.name} (ID: {employee.person_id})"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.append([])
    
    headers = ['Date', 'Day', 'First In', 'Last Out', 'Duration', 'Status']
    ws.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    _, days_in_month = calendar.monthrange(selected_year, selected_month)
    
    full_days = 0
    half_days = 0
    leave_days = 0
    paid_leave_days = 0
    late_arrivals = 0
    holidays_count = 0
    
    # Default shift times (can be customized per employee if ShiftHistory is used)
    emp_shift_start = time(9, 0)  # 9:00 AM
    emp_shift_end = time(18, 0)   # 6:00 PM

    month_start = datetime.date(selected_year, selected_month, 1)
    month_end = datetime.date(selected_year, selected_month, days_in_month)

    approved_leaves = LeaveRequest.objects.filter(
        employee=employee,
        status='approved',
        start_date__lte=month_end,
        end_date__gte=month_start
    )
    approved_leave_dates = set()
    for leave in approved_leaves:
        start = max(leave.start_date, month_start)
        end = min(leave.end_date, month_end)
        curr = start
        while curr <= end:
            approved_leave_dates.add(curr)
            curr += datetime.timedelta(days=1)
    
    day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    for day in range(1, days_in_month + 1):
        date = datetime.date(selected_year, selected_month, day)
        weekday = date.weekday()
        day_name = day_names[weekday]
        is_sunday = weekday == 6
        is_holiday = date in holiday_dates
        is_saturday = weekday == 5
        
        record = records_dict.get(date)
        
        if is_sunday or is_holiday:
            holidays_count += 1
            status = "Holiday"
            first_in = "-"
            last_out = "-"
            duration = "-"
            fill = sunday_fill if is_sunday else holiday_fill
        elif record:
            first_in = record.first_in.strftime("%H:%M") if record.first_in else "-"
            last_out = record.last_out.strftime("%H:%M") if record.last_out else "-"
            duration = str(record.work_duration) if record.work_duration else "-"
            
            total_secs = record.work_duration.total_seconds() if record.work_duration else 0
            is_late = record.first_in and record.first_in > emp_shift_start
            arrived_after_noon = record.first_in and record.first_in.hour >= 12
            
            if is_saturday:
                sat_shift_end = time(emp_shift_start.hour + 4, emp_shift_start.minute)
                left_early = record.last_out and (
                    record.last_out.hour < sat_shift_end.hour or 
                    (record.last_out.hour == sat_shift_end.hour and record.last_out.minute < sat_shift_end.minute)
                )
            else:
                left_early = record.last_out and (
                    record.last_out.hour < emp_shift_end.hour or 
                    (record.last_out.hour == emp_shift_end.hour and record.last_out.minute < emp_shift_end.minute)
                )
            
            is_half_day = arrived_after_noon or left_early
            
            if total_secs == 0:
                if date in approved_leave_dates and not is_sunday and not is_holiday:
                    status = "Paid Leave"
                    fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Blue-ish
                    paid_leave_days += 1
                else:
                    status = "Leave"
                    fill = red_fill
                    leave_days += 1
            elif is_half_day:
                status = "Half Day"
                fill = yellow_fill
                half_days += 1
                if is_late:
                    late_arrivals += 1
            elif is_late:
                status = "Late"
                fill = yellow_fill
                full_days += 1
                late_arrivals += 1
            else:
                status = "Present"
                fill = green_fill
                full_days += 1
        else:
            first_in = "-"
            last_out = "-"
            duration = "-"
            if date <= datetime.date.today():
                if date in approved_leave_dates and not is_sunday and not is_holiday:
                    status = "Paid Leave"
                    fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                    paid_leave_days += 1
                else:
                    status = "Leave"
                    fill = red_fill
                    leave_days += 1
            else:
                status = "-"
                fill = None
        
        row = [
            date.strftime("%Y-%m-%d"),
            day_name,
            first_in,
            last_out,
            duration,
            status
        ]
        ws.append(row)
        
        current_row = ws.max_row
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if fill:
                cell.fill = fill
    
    ws.append([])
    ws.append([])
    
    summary_row = ws.max_row + 1
    ws.merge_cells(f'A{summary_row}:F{summary_row}')
    ws.cell(row=summary_row, column=1).value = "Monthly Summary"
    ws.cell(row=summary_row, column=1).font = title_font
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center')
    
    summary_data = [
        ("Full Days", full_days),
        ("Half Days", half_days),
        ("Leave Days", leave_days),
        ("Late Arrivals", late_arrivals),
        ("Holidays/Sundays", holidays_count),
        ("Working Days", full_days + (half_days * 0.5) + holidays_count + paid_leave_days),
        ("Paid Leave Days", paid_leave_days)
    ]
    
    for label, value in summary_data:
        ws.append([label, value])
        current_row = ws.max_row
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = "".join(c for c in employee.name if c.isalnum() or c in (' ', '-', '_')).strip()
    safe_name = safe_name.replace(' ', '_')
    filename = f"{safe_name}_Attendance_{selected_year}_{selected_month:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Content-Length'] = len(buffer.getvalue())
    
    return response


@login_required
def download_remote_report(request):
    """Generate and download XLSX report for remote team."""
    now = datetime.datetime.now()
    try:
        selected_month = int(request.GET.get('month', now.month))
        selected_year = int(request.GET.get('year', now.year))
    except ValueError:
        selected_month = now.month
        selected_year = now.year
    
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[selected_month]
    
    _, days_in_month = calendar.monthrange(selected_year, selected_month)
    sundays_count = 0
    for day in range(1, days_in_month + 1):
        date = datetime.date(selected_year, selected_month, day)
        if date.weekday() == 6:
            sundays_count += 1
    
    show_inactive = request.GET.get('show_inactive', '') == '1'
    
    summaries = RemoteMonthlySummary.objects.filter(
        year=selected_year,
        month=selected_month
    ).select_related('employee')
    
    if not show_inactive:
        summaries = summaries.filter(employee__is_active=True)
    
    summaries = summaries.order_by('employee__name')
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {selected_year}"
    
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Remote Team Attendance Report - {month_name} {selected_year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.append([])
    
    holidays_in_month = Holiday.objects.filter(
        date__year=selected_year,
        date__month=selected_month
    ).count()
    total_holidays = sundays_count + holidays_in_month
    
    headers = ['Employee Name', 'Full Days', 'Half Days', 'Leave Days', 'Holidays', 'Working Days']
    ws.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for summary in summaries:
        half_days = summary.half_days
        full_days = max(0, summary.present_days - half_days)
        leave_days = summary.absent_days
        working_days_total = full_days + (half_days * 0.5) + total_holidays
        
        row = [
            summary.employee.name,
            full_days,
            half_days,
            leave_days,
            total_holidays,
            working_days_total
        ]
        ws.append(row)
        
        current_row = ws.max_row
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            if col > 1:
                cell.alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Remote_Attendance_Report_{selected_year}_{selected_month:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Content-Length'] = len(buffer.getvalue())
    
    return response


@login_required
def download_remote_employee_report(request, employee_id):
    """Generate and download XLSX report for a single remote employee for the selected month."""
    employee = get_object_or_404(RemoteEmployee, id=employee_id)
    
    now = datetime.datetime.now()
    try:
        selected_month = int(request.GET.get('month', now.month))
        selected_year = int(request.GET.get('year', now.year))
    except ValueError:
        selected_month = now.month
        selected_year = now.year
    
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[selected_month]
    
    holiday_dates = set(Holiday.objects.filter(
        date__year=selected_year,
        date__month=selected_month
    ).values_list('date', flat=True))
    
    records = RemoteCallRecord.objects.filter(
        employee=employee,
        date__year=selected_year,
        date__month=selected_month
    ).order_by('date')
    
    records_dict = {r.date: r for r in records}
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{employee.name[:20]}"
    
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    holiday_fill = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    red_fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Remote Call Statistics - {month_name} {selected_year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Employee: {employee.name} (Extension: {employee.extension_id})"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.append([])
    
    headers = ['Date', 'Day', 'Answered Calls', 'Talk Duration', 'Status']
    ws.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    _, days_in_month = calendar.monthrange(selected_year, selected_month)
    
    present_days = 0
    half_days = 0
    absent_days = 0
    holidays_count = 0
    total_calls = 0
    total_talk_minutes = 0
    
    day_names = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    for day in range(1, days_in_month + 1):
        date = datetime.date(selected_year, selected_month, day)
        weekday = date.weekday()
        day_name = day_names[weekday]
        is_sunday = weekday == 6
        is_holiday = date in holiday_dates
        
        record = records_dict.get(date)
        
        if is_sunday or is_holiday:
            holidays_count += 1
            status = "Holiday"
            answered_calls = "-"
            talk_duration = "-"
            fill = holiday_fill
        elif record:
            answered_calls = record.answered_calls or 0
            talk_minutes = int(record.total_talk_duration.total_seconds() / 60) if record.total_talk_duration else 0
            talk_duration = f"{talk_minutes} min"
            total_calls += answered_calls
            total_talk_minutes += talk_minutes
            
            if record.attendance_status == 'present':
                status = "Present"
                fill = green_fill
                present_days += 1
            elif record.attendance_status == 'half_day':
                status = "Half Day"
                fill = yellow_fill
                half_days += 1
            else:
                status = "Absent"
                fill = red_fill
                absent_days += 1
        else:
            answered_calls = "-"
            talk_duration = "-"
            if date <= datetime.date.today():
                status = "No Data"
                fill = red_fill
                absent_days += 1
            else:
                status = "-"
                fill = None
        
        row = [
            date.strftime("%Y-%m-%d"),
            day_name,
            answered_calls,
            talk_duration,
            status
        ]
        ws.append(row)
        
        current_row = ws.max_row
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if fill:
                cell.fill = fill
    
    ws.append([])
    ws.append([])
    
    summary_row = ws.max_row + 1
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    ws.cell(row=summary_row, column=1).value = "Monthly Summary"
    ws.cell(row=summary_row, column=1).font = title_font
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center')
    
    summary_data = [
        ("Present Days", present_days),
        ("Half Days", half_days),
        ("Absent Days", absent_days),
        ("Holidays/Sundays", holidays_count),
        ("Total Calls Answered", total_calls),
        ("Total Talk Time", f"{total_talk_minutes} min"),
        ("Working Days", present_days + (half_days * 0.5) + holidays_count)
    ]
    
    for label, value in summary_data:
        ws.append([label, value])
        current_row = ws.max_row
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).border = thin_border
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = "".join(c for c in employee.name if c.isalnum() or c in (' ', '-', '_')).strip()
    safe_name = safe_name.replace(' ', '_')
    filename = f"{safe_name}_Remote_Stats_{selected_year}_{selected_month:02d}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    response['Content-Length'] = len(buffer.getvalue())
    
    return response
